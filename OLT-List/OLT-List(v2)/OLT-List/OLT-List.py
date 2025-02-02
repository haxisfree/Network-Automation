from netmiko import ConnectHandler
from getpass import getpass
from getpass import getuser
import regex as re
import time
import openpyxl
import os
from colorama import init, Fore, Back, Style
import datetime
#################################

st = time.time()


init()

list_of_description_filename = "List Of Description.txt"
relative_path = os.path.dirname(__file__)
list_of_description_filename_absolute_path = relative_path + "\\" + list_of_description_filename
pppoe7_subif = []
pppoe4_subif = []



# DNS List
tehran_dns_list = ['pppoe7.dc2.rou', 'pppoe4.col.rou'] #, 'pppoe4.dc2.rou']
#pardis_dns_list = ['pppoe2.dc2.rou']

# Deficiency Lists
pppoe7_list_deficiency = []
pppoe4_list_deficiency = []


sub_int_list = []
prd_sub_int_list = []









class File:

    def __init__(self):
        pass

    def Excel_File(self):

        # Get Excel Filename
        excel_file_path = input("Please specify the Excel file path. Otherwise to use default list press ENTER: ")
        if excel_file_path == "":
            for filename in os.listdir(relative_path):
                if "xlsx" in filename:
                    excel_file_name = re.findall(r".*\.xlsx$", filename)
                    excel_file_path = relative_path + "\\" + excel_file_name[0]
                    print(excel_file_path)
                else:
                    pass
        else:
            excel_file_path = re.sub('"', '', excel_file_path)
            print (excel_file_path)
            # pass
  
        
        # Load Excel
        wrkbk = openpyxl.load_workbook(excel_file_path) 
        sh = wrkbk.active
    
        # List Of Row Numbers
        row_numbers = []

        # Get Number Of Rows
        for i in range (1, sh.max_row+1):
            row_numbers.append(i)

        # Regex Pattern To Seprate Bras Sub Interfaces Lists
        for row in row_numbers[1:]:
            cell_obj = sh.cell(row,1)
            
            if "C1002.07" in cell_obj.value:
                sub_int = re.findall(r"\d*\/\d*\/\d*\.\d*", cell_obj.value)
                pppoe7_subif.append(sub_int[0])
                
            elif "C1002.04" in cell_obj.value:
                sub_int = re.findall(r"\d*\/\d*\/\d*\.\d*", cell_obj.value)
                pppoe4_subif.append(sub_int[0])
            else:
                continue
        return pppoe7_subif, pppoe4_subif



    def existing_file_question(self):
        for filename in os.listdir(relative_path):
            if "List Of Description.txt" in filename:
                overwrite = input("There is another file with name \"List Of Description\". Do you want to overwrite it?(y/n) ")
                while True:
                    if overwrite == "":
                        overwrite = input("\n\nAnswer by yes or no:\nThere is another file with name \"List Of Description\". Do you want to overwrite it?(y/n) ")

                    elif overwrite in "y,yes,YES,Yes,yES,غثس,غ,":
                        with open("List Of Description.txt", "w") as f:
                            f.close()
                        break
                    
                    elif overwrite in "no,No,NO,n,N":
                        with open(filename, 'a') as f:
                            # for item in sub_int_list:
                            #     f.write(f"{item}\n")
                            #     f.write(f"-" * 50 + "\n")
                                f.close()
                        break
                    
                    else:
                        overwrite = input("\n\nAnswer by yes or no:\nThere is another file with name \"List Of Description\". Do you want to overwrite it?(y/n) ")
            
            else:
                pass
             

    def __str__(self):
        return f"This Class, returns all objects in the input file with list of readable object (without any space)"



class TextToTable:

    def __init__(self, textlist):
        self.textlist = textlist


    def Line_Split_Strip(self):
        linesplited_item = self.textlist.splitlines()
        Items = []
        for item in linesplited_item:
            striped_item = item.strip()
            columnsplited_item = re.split(r'\s+', striped_item)
            Items.append(columnsplited_item)
            # print(Items)       
        
        return Items


excel_file_ins = File()
excel_file_ins.existing_file_question()
excel_file_output = excel_file_ins.Excel_File()


class ConnectionTest:
    def __init__(self):
        pass
        
    def connection(self):
        count = 0
        while count < 3 :
            try:
                self.Username = input('Please enter your username:')
                self.Password = getpass('Please enter your password:')
                info = { 
                "device_type": "cisco_ios",
                "host": "l22.dc1.swi",
                "username": self.Username,
                "password": self.Password,
                }        
                with ConnectHandler(**info) as net_connect:
                    net_connect.disconnect
                
                count = 3
                return self.Username, self.Password
            
            except Exception as error:
                pass
                print("An error occurred:", error) # An error occurred: name 'x' is not defined
                count += 1

    




class Shows:
    
    def __init__(self):
        # Show command that we execute.
            self.tehran_command = "show int desc | i .TEH.OLT"
            self.pardis_command = "show int desc | i .PRD.OLT"



    """
    The Connection
    Regex pattent to match the needed information
    Comparison
    Write to file
    """
    def Connection(self):
        # Start Of Connecten
        # Send Command
        # with ConnectHandler(**tehran_connection_info) as net_connect:
        confed_sub_int_list_dic = {}
        confed_prd_sub_int_list_dic = {}
        show_output = net_connect.send_command(self.tehran_command)
        pardis_show_output = net_connect.send_command(self.pardis_command)

        
        # Split Output Into List Of Lines
        linesplited_show_output_ins = TextToTable(show_output)
        linesplited_show_output = linesplited_show_output_ins.Line_Split_Strip()
        prd_linesplited_show_output_ins = TextToTable(pardis_show_output)
        prd_linesplited_show_output = prd_linesplited_show_output_ins.Line_Split_Strip()
        
        # Iteration Through Splited Lines
        for item in linesplited_show_output:

            # Only Up/Up Sub Interfaces
            if item[1] == "up" == item[2] == "up":
                sub_int = re.findall(r"\d*\/\d*\/\d*\.\d*", item[0])
                sub_int_list.append(sub_int[0])
                confed_sub_int_list_dic[sub_int[0]] = item[3]
            else:
                continue
        
        
        
        # Iteration Through Splited Lines
        for item in prd_linesplited_show_output:

            # Only Up/Up Sub Interfaces
            if item[1] == "up" == item[2] == "up":
                sub_int = re.findall(r"\d*\/\d*\/\d*\.\d*", item[0])
                prd_sub_int_list.append(sub_int[0])
                confed_prd_sub_int_list_dic[sub_int[0]] = item[3]
            else:
                continue
   
        #print (sub_int_list_dic)
        return confed_sub_int_list_dic, confed_prd_sub_int_list_dic





class Compare:

    def __init__(self, confed_sub_int_list_dic, pppoe7_subif, pppoe4_subif ):
        self.confed_sub_int_list_dic = confed_sub_int_list_dic
        self.pppoe7_subif = pppoe7_subif
        self.pppoe4_subif = pppoe4_subif
    
    
    def compare(self):

        # Comparison Between List's Sub Interfaces and Device Configured Sub Interfaces
        if i == "pppoe7.dc2.rou":
            #print (sub_int_list_dic)
            for sub_int in self.confed_sub_int_list_dic:
                if sub_int in self.pppoe7_subif: #or line0 in pardis_show_output:
                    # Remove To Clearify What Remains
                    self.pppoe7_subif.remove(sub_int)
                else:
                    # Sub Interfaces That Are Not In List But Are Configured In Device
                    pppoe7_list_deficiency.append(sub_int)

        
        # Loop == 1 Is For Bras4
        elif i == "pppoe4.col.rou":
            for sub_int in self.confed_sub_int_list_dic:
                if sub_int in self.pppoe4_subif:# or line0 in self.pardis_show_output:
                    # Remove To Clearify What Remains
                    self.pppoe4_subif.remove(sub_int)
                else:
                    # Sub Interfaces That Are Not In List But Are Configured In Device
                    pppoe4_list_deficiency.append(sub_int)
                    
        else:
            pass

        return self.pppoe7_subif , self.pppoe4_subif















class Print:

    def __init__(self, pppoe4_subif, pppoe7_subif, pppoe4_list_deficiency, pppoe7_list_deficiency,confed_sub_int_list_dic,confed_prd_sub_int_list_dic):
        
        self.pppoe4_subif = pppoe4_subif
        self.pppoe7_subif = pppoe7_subif
        self.pppoe4_list_deficiency = pppoe4_list_deficiency
        self.pppoe7_list_deficiency = pppoe7_list_deficiency
        self.confed_prd_sub_int_list_dic = confed_prd_sub_int_list_dic
        self.confed_sub_int_list_dic = confed_sub_int_list_dic

    def file_print(self):
        with open('List Of Description.txt', 'a') as f:
            f.write(f"-" * 30 + str(i) + "-" * 30 + "\n")
            
            if i == "pppoe7.dc2.rou":
                f.write("The configured sub-interfaces (they DO NOT exist in excel):")
                for item in self.pppoe7_list_deficiency:
                    if item in self.confed_prd_sub_int_list_dic.keys():
                        f.write(item + "  " + self.confed_prd_sub_int_list_dic[item] )
                    else:
                        f.write(item + "  " + self.confed_sub_int_list_dic[item])
                f.write("The unconfigured sub-interfaces (they exist in excel):")
                for item in self.pppoe7_subif:
                    f.write(item)





def Print(confed_sub_int_list_dic,confed_prd_sub_int_list_dic,i):

    # Write Sub Interface That Are Configured In Device Into A List 
    with open('List Of Description.txt', 'a') as f:
        # f.write(f"-" * 30 + str(i) + "-" * 30 + "\n")
        # for line1 in confed_sub_int_list_dic:
        #     f.write(f"{line1}   {confed_sub_int_list_dic[line1]}\n")
        # for line1 in confed_prd_sub_int_list_dic:
        #     f.write(f"{line1}   {confed_prd_sub_int_list_dic[line1]}\n")
        
    
        
        if i == "pppoe7.dc2.rou":
            f.write("="*40 + str(datetime.datetime.now()) + "="*40 + "\n")
            print ("-" * 50)
            print ("\n" * 5)
            print ("-" * 50 + Fore.RED + Back.WHITE + "BRAS 7" + Style.RESET_ALL + "-" * 50 )
            f.write("-" * 50 + "BRAS 7" + "-" * 50 + "\n" )
            print ("The configured sub-interfaces (they DO NOT exist in excel):")
            f.write("The configured sub-interfaces (they DO NOT exist in excel):" + "\n")
            for line0 in pppoe7_list_deficiency:
                if line0 in confed_prd_sub_int_list_dic.keys():
                    print(line0 + "  " + confed_prd_sub_int_list_dic[line0] )
                    f.write(line0 + "  " + confed_prd_sub_int_list_dic[line0] + "\n")
                else:
                    print(line0 + "  " + confed_sub_int_list_dic[line0])
                    f.write(line0 + "  " + confed_sub_int_list_dic[line0] + "\n")
                #print (line0 + "  " + sub_int_list_dic[line0])
            for si1 in pppoe7_subif:
                if si1 in confed_prd_sub_int_list_dic.keys():
                    print(si1 + "  " + confed_prd_sub_int_list_dic[si1] )
                    f.write(si1 + "  " + confed_prd_sub_int_list_dic[si1] + "\n")
                else:
                    pass
            print (Fore.RED + "." * 10 + Style.RESET_ALL)
            f.write ("." * 10 + "\n")
            print ("The unconfigured sub-interfaces (they exist in excel):")
            f.write ("The unconfigured sub-interfaces (they exist in excel):"+ "\n")
            for si1 in pppoe7_subif:
                if si1 in confed_prd_sub_int_list_dic.keys():
                    pass
                else:
                    print(si1)
                    f.write(si1 + "\n")
            print("\t")
            f.write("\n")
        
        # Loop == 1 Is For Bras4
        elif i == "pppoe4.col.rou":
            #print ("-" * 50)
            #print ("\n" * 1)
            print ("-" * 50 + Fore.RED + Back.WHITE + "BRAS 4" + Style.RESET_ALL + "-" * 50 )
            f.write ("-" * 50 + "BRAS 4" + "-" * 50 + "\n")
            print ("The configured sub-interfaces (they DO NOT exist in excel):")
            f.write ("The configured sub-interfaces (they DO NOT exist in excel):"+ "\n")
            for line0 in pppoe4_list_deficiency:
                if line0 in confed_prd_sub_int_list_dic.keys():
                    print(line0 + "  " + confed_prd_sub_int_list_dic[line0] )
                    f.write(line0 + "  " + confed_prd_sub_int_list_dic[line0] + "\n")
                else:
                    print(line0 + "  " + confed_sub_int_list_dic[line0])
                    f.write(line0 + "  " + confed_sub_int_list_dic[line0]+ "\n")
            
            for si1 in pppoe4_subif:
                if si1 in confed_prd_sub_int_list_dic.keys():
                    print(si1 + "  " + confed_prd_sub_int_list_dic[si1] )
                    f.write(si1 + "  " + confed_prd_sub_int_list_dic[si1]+ "\n" )
                else:
                    pass
            print (Fore.RED + "." * 10 + Style.RESET_ALL)
            f.write ("." * 10 + "\n")
            print ("The unconfigured sub-interfaces (they exist in excel):")
            f.write ("The unconfigured sub-interfaces (they exist in excel):"+ "\n")
            for si1 in pppoe4_subif:
                if si1 in confed_prd_sub_int_list_dic.keys():
                    pass
                else:
                    print(si1)
                    f.write(si1+ "\n")
            f.write("\n\n\n")
        else:
            pass
        f.close()
    

















# User/Pass

try:
    connection_test = ConnectionTest()
    Username,Password = connection_test.connection()
    for i in tehran_dns_list :
        tehran_connection_info = { 
        "device_type": "cisco_ios",
        "host": str(i),
        "username": Username,
        "password": Password,
        }

        with ConnectHandler(**tehran_connection_info) as net_connect:
            shows_ins = Shows()
            shows_result = shows_ins.Connection()
            # print(y[0])
            deficiencies_ins = Compare(shows_result[0], excel_file_output[0], excel_file_output[1])
            list_deficiencies = deficiencies_ins.compare()
            Print(shows_result[0],shows_result[1],i)

except:
    print("Something went wrong!!!")



et = time.time()
elapsed_time = et - st


# Done Message
print ("\n" * 2)
print("!!!!! Its Done !!!!!")


print ("\nScript name : OLT-List")
print ("Time to run script : ", end = " ")
print (elapsed_time)

print("\n")
print (Style.BRIGHT+Fore.RED+"G"+Style.RESET_ALL+Style.BRIGHT+Fore.CYAN+"r"+Style.RESET_ALL+Style.BRIGHT+Fore.YELLOW+"o"+Style.RESET_ALL+Style.BRIGHT+Fore.GREEN+"u"+Style.RESET_ALL+Style.BRIGHT+Fore.WHITE+"p "+Style.RESET_ALL+Style.BRIGHT+Fore.MAGENTA+"A"+Style.RESET_ALL)
print (Style.BRIGHT+Fore.RED+"G"+Style.RESET_ALL+Style.BRIGHT+Fore.CYAN+"r"+Style.RESET_ALL+Style.BRIGHT+Fore.YELLOW+"o"+Style.RESET_ALL+Style.BRIGHT+Fore.GREEN+"u"+Style.RESET_ALL+Style.BRIGHT+Fore.WHITE+"p "+Style.RESET_ALL+Style.BRIGHT+Fore.MAGENTA+"A"+Style.RESET_ALL)
print (Style.BRIGHT+Fore.RED+"G"+Style.RESET_ALL+Style.BRIGHT+Fore.CYAN+"r"+Style.RESET_ALL+Style.BRIGHT+Fore.YELLOW+"o"+Style.RESET_ALL+Style.BRIGHT+Fore.GREEN+"u"+Style.RESET_ALL+Style.BRIGHT+Fore.WHITE+"p "+Style.RESET_ALL+Style.BRIGHT+Fore.MAGENTA+"A"+Style.RESET_ALL)
print (Style.BRIGHT+Fore.RED+"G"+Style.RESET_ALL+Style.BRIGHT+Fore.CYAN+"r"+Style.RESET_ALL+Style.BRIGHT+Fore.YELLOW+"o"+Style.RESET_ALL+Style.BRIGHT+Fore.GREEN+"u"+Style.RESET_ALL+Style.BRIGHT+Fore.WHITE+"p "+Style.RESET_ALL+Style.BRIGHT+Fore.MAGENTA+"A"+Style.RESET_ALL)
print (Style.BRIGHT+Fore.RED+"G"+Style.RESET_ALL+Style.BRIGHT+Fore.CYAN+"r"+Style.RESET_ALL+Style.BRIGHT+Fore.YELLOW+"o"+Style.RESET_ALL+Style.BRIGHT+Fore.GREEN+"u"+Style.RESET_ALL+Style.BRIGHT+Fore.WHITE+"p "+Style.RESET_ALL+Style.BRIGHT+Fore.MAGENTA+"A"+Style.RESET_ALL)

time.sleep(600)